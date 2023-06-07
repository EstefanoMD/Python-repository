from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options 
import openpyxl

# Carregue a planilha
arquivo_planilha = openpyxl.load_workbook(r'C:\Users\estef\Documents\Python\Projeto_quality_automatizacao\Planilha teste.xlsx')

# Selecione a planilha
planilha = arquivo_planilha['Planilha']

# Obtenha os dados da planilha

nota_fiscal = []
for valor in planilha.iter_rows(min_row=3, min_col=12, values_only=True):
    inteiro = int(valor[0])
    nota_fiscal.append(inteiro)
print(nota_fiscal)

data = []
for valor in planilha.iter_rows(min_row=3, min_col=5, values_only=True):
    string = str(valor[0])
    data.append(string)
print(data)

rota_origem = []
for valor in planilha.iter_rows(min_row=3, min_col=7, values_only=True):
    rota_origem.append(valor[0])
print(rota_origem)

rota_destino = []
for valor in planilha.iter_rows(min_row=3, min_col=8, values_only=True):
    rota_destino.append(valor[0])
print(rota_destino)

cpf = []
for valor in planilha.iter_rows(min_row=3, min_col=17, values_only=True):
    string = str(valor[0])
    cpf.append(string)
print(cpf)

options = Options()
options.add_experimental_option("detach", True)
driver = webdriver.Chrome(r'C:\Users\estef\Documents\WebDrivers\chromedriver_win32\chromedriver',chrome_options=options)
driver.get('https://docs.google.com/forms/d/e/1FAIpQLSduUoUhCBzuMXoA8XmLTAjWZ5KUQW5j5sgGYzVwLUIbQhi_kA/viewform?vc=0&c=0&w=1&flr=0')

# Preencha cada campo do formulário com os dados da planilha
driver.find_element(By.CLASS_NAME, "whsOnd zHQkBf").send_keys(nota_fiscal)
# driver.find_element(By.NAME, "whsOnd zHQkBf").send_keys(data)
# # Adicione mais linhas de código semelhantes para preencher os outros campos

# # Envie o formulário
# driver.find_element(By.XPATH, '//button[contains(text(), "Enviar")]').click()

# # Aguarde um tempo antes de prosseguir para o próximo conjunto de dados
# driver.implicitly_wait(3)